from pathlib import Path
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

def autosize_and_style(ws):
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for column_cells in ws.columns:
        max_len = 0
        col_idx = column_cells[0].column
        for cell in column_cells[:200]:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 10), 40)

def write_workbook(output_path: Path, participant: str, streams: dict,
                   overview_rows: list[dict], summary_rows: list[dict]):
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        overview = pd.DataFrame(overview_rows)
        overview.to_excel(writer, sheet_name="Overview", index=False)

        for stream, df in streams.items():
            safe_name = stream[:31]
            df.to_excel(writer, sheet_name=safe_name, index=False)

        pd.DataFrame(summary_rows).to_excel(
            writer, sheet_name="Summary", index=False
        )

        wb = writer.book
        for ws in wb.worksheets:
            autosize_and_style(ws)

        # Make Overview and Summary visually useful.
        for name in ("Overview", "Summary"):
            ws = wb[name]
            ws.sheet_view.showGridLines = False

        wb.properties.title = f"Cognitive Fatigue Data - {participant}"
        wb.properties.subject = "Participant data consolidation"
        wb.properties.creator = "Cognitive Fatigue Data Consolidator"
